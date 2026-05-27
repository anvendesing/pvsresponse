"""
Matches DB products to product images from the PVS Excel master sheet.

Strategy:
  - Excel col D = clean product name, col B = image ref (e.g. "I97")
  - Images in imagespvs1x_Upload/ are named I971PrakruthiVanam22122021.jpg
    (pattern: <ref> + optional "1" + "PrakruthiVanam" + date + .jpg)
  - DB Product.sku differs from Excel Code column — match by normalised name
  - Normalise: lowercase, strip size/weight tokens, remove non-alnum chars,
    then check substring containment in either direction.

Run from repo root:  python scripts/import-product-images.py
"""

import os, re, shutil, sqlite3
from pathlib import Path
import openpyxl

REPO     = Path(__file__).resolve().parent.parent
EXCEL    = REPO / "ItemMaster_PVS_05102023.xlsx"
IMG_SRC  = REPO / "imagespvs1x_Upload"
IMG_DEST = REPO / "backend" / "uploads" / "products"
DB_PATH  = REPO / "backend" / "prisma" / "dev.db"

SIZE_PAT = re.compile(
    r'\b\d+\s*(g|gm|gms|kg|ml|l|ltr|lt|litre|pcs|pc|nos|no|set|pack|'
    r'sachet|pouch|bottle|jar|box)\b', re.IGNORECASE
)

def flatten(s: str) -> str:
    s = SIZE_PAT.sub(' ', s)
    s = s.lower()
    s = re.sub(r'[^a-z0-9]', '', s)
    return s


# ── 1. Read Excel ─────────────────────────────────────────────────────────────
print("Reading Excel ...")
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb.active

excel_rows: list[tuple[str, str, str]] = []   # (raw_name, flat_name, ref)
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 4).value
    ref  = ws.cell(r, 2).value
    if not name or not ref:
        continue
    ref_s = str(ref).strip()
    if ref_s in ("#N/A", ""):
        continue
    raw = str(name).strip()
    excel_rows.append((raw, flatten(raw), ref_s))

print(f"  {len(excel_rows)} Excel rows with valid image ref")

# ── 2. Build ref → filename (search the folder for each ref) ─────────────────
print("Scanning image folder ...")
all_img_files = os.listdir(IMG_SRC)

def find_file_for_ref(ref: str) -> str | None:
    """Return first image filename in IMG_SRC whose name starts with
    <ref> optionally followed by a single '1' before 'PrakruthiVanam'."""
    pattern = re.compile(
        re.escape(ref) + r'1?PrakruthiVanam', re.IGNORECASE
    )
    for fname in all_img_files:
        if pattern.match(fname):
            return fname
    return None

# Pre-build unique-ref set from Excel so we don't re-scan for duplicates
unique_refs = {ref for _, _, ref in excel_rows}
ref_to_file: dict[str, str] = {}
for ref in unique_refs:
    f = find_file_for_ref(ref)
    if f:
        ref_to_file[ref.upper()] = f

print(f"  {len(ref_to_file)} image refs resolved to files")

# ── 3. Read DB products ───────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
cur  = conn.cursor()
cur.execute("SELECT id, sku, name FROM Product ORDER BY name")
db_products = cur.fetchall()
conn.close()

print(f"  {len(db_products)} products in DB")

# ── 4. Match DB product → Excel ref (name-based) ─────────────────────────────
def best_ref(db_flat: str) -> str | None:
    # Pass 1: exact flat match
    for _, ex_flat, ref in excel_rows:
        if db_flat == ex_flat:
            return ref
    # Pass 2: substring containment (both >= 5 chars to avoid false positives)
    if len(db_flat) >= 5:
        for _, ex_flat, ref in excel_rows:
            if len(ex_flat) >= 5 and (db_flat in ex_flat or ex_flat in db_flat):
                return ref
    return None

# ── 5. Copy images and collect DB updates ────────────────────────────────────
IMG_DEST.mkdir(parents=True, exist_ok=True)

updates:   list[tuple[str, str]] = []   # (imageUrl, product_id)
no_ref:    list[str] = []
no_file:   list[str] = []
newly_copied: list[str] = []

for pid, sku, name in db_products:
    db_flat = flatten(name)
    ref = best_ref(db_flat)
    if ref is None:
        no_ref.append(f"{sku} | {name}")
        continue

    ref_upper = ref.upper()
    if ref_upper not in ref_to_file:
        no_file.append(f"{sku} | {name} -> ref={ref}")
        continue

    src_name  = ref_to_file[ref_upper]
    ext       = Path(src_name).suffix.lower()
    dest_name = f"{ref_upper}{ext}"
    dest_path = IMG_DEST / dest_name

    if not dest_path.exists():
        shutil.copy2(IMG_SRC / src_name, dest_path)
        newly_copied.append(dest_name)

    updates.append((f"/uploads/products/{dest_name}", pid))

print(f"\nMatched: {len(updates)}  |  No ref in Excel: {len(no_ref)}  |  Ref has no image file: {len(no_file)}")
print(f"New images copied to uploads/products/: {len(newly_copied)}")

# ── 6. Update DB ─────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
cur  = conn.cursor()
for url, pid in updates:
    cur.execute("UPDATE Product SET imageUrl = ? WHERE id = ?", (url, pid))
conn.commit()
conn.close()

print(f"DB updated: {len(updates)} products now have imageUrl")

# ── 7. Report ─────────────────────────────────────────────────────────────────
if no_ref:
    print(f"\nProducts with no name match in Excel ({len(no_ref)}) - these have no source image:")
    for s in no_ref[:30]:
        print("  ", s)
if no_file:
    print(f"\nExcel refs with no image file ({len(no_file)}):")
    for s in no_file[:20]:
        print("  ", s)

print("\nDone.")
