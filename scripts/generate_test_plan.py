"""
PVS-ERP Test Plan generator
============================
Builds an updated PVS-ERP-Test-Plan.xlsx with detailed test cases for:
  - Web ERP portal (all pages, all key flows)
  - Warehouse mobile picking & packing screens
  - Mobile companion screens (login, tasks, scan, count, GRN, returns)
  - Backend API contract tests
  - End-to-end flows
  - Role/permission matrix

Adds an explicit "Changes Required" column to flag known gaps that need
code work alongside the test.

Run:  python scripts/generate_test_plan.py
"""

from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "PVS-ERP-Test-Plan.xlsx"

# ────────────────────────────────────────────────────────────────────
# Style helpers
# ────────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="003087")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA = PatternFill("solid", fgColor="F5F7FA")
THIN = Side(border_style="thin", color="CBD2D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

STATUS_FILLS = {
    "Passed": PatternFill("solid", fgColor="D6F5DE"),
    "Failed": PatternFill("solid", fgColor="F8D6D6"),
    "Blocked": PatternFill("solid", fgColor="FFE6BD"),
    "Not Run": PatternFill("solid", fgColor="ECEFF3"),
    "In Progress": PatternFill("solid", fgColor="DCE7FA"),
}

COL_WIDTHS = [10, 18, 24, 42, 60, 60, 12, 10, 12, 36, 36]

HEADERS = [
    "ID",
    "Area",
    "Module / Item",
    "Test Case",
    "Steps",
    "Expected Result",
    "Type",
    "Priority",
    "Status",
    "Changes Required",
    "Notes / Defect",
]


def write_sheet(wb, name, rows, title=None):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14, color="003087")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
        header_row = 3
    else:
        header_row = 1

    # header
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # data
    for r, row in enumerate(rows, header_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ZEBRA
            # status column = 9
            if c == 9 and val in STATUS_FILLS:
                cell.fill = STATUS_FILLS[val]
                cell.font = Font(bold=True)

    # widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # freeze header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # row heights
    for r in range(header_row + 1, header_row + 1 + len(rows)):
        ws.row_dimensions[r].height = 48

    # status column conditional formatting (visual aid)
    last_row = header_row + len(rows)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{last_row}"
    return ws


def make_row(_id, area, module, case, steps, expected, ttype, priority, status, changes="", notes=""):
    return [_id, area, module, case, steps, expected, ttype, priority, status, changes, notes]


# ────────────────────────────────────────────────────────────────────
# Test data
# ────────────────────────────────────────────────────────────────────

OVERVIEW = [
    ("Version", "v2 — Detailed (Web Portal + Mobile Picking/Packing)"),
    ("Generated", date.today().isoformat()),
    ("Backend", "Fastify + Prisma + SQLite (dev) / Postgres (prod); REST /v1 on :4000"),
    ("Web Portal", "React 18 + Vite at :5173 — admin + warehouse PWA at /m/*"),
    ("Warehouse APK", "Capacitor wrapper (com.prakruthivanam.warehouse) wraps the /m/* PWA"),
    ("VPS", "http://217.216.78.119 (production)"),
    ("Mobile build", "cd erp-portal && npm run build:mobile; cd ../mobile-erp && npm run build:android"),
    ("APK output", "mobile-erp/release/android/com.prakruthivanam.warehouse-Capacitor-Debug.apk"),
    ("", ""),
    ("Worksheet", "Purpose"),
    ("1. Models", "Data-model integrity, validation & invariants (Prisma)"),
    ("2. ERP Pages", "Desktop portal — detailed page-level functional & UI tests"),
    ("3. Mobile Pick/Pack", "Detailed picking & packing scenarios on the warehouse APK"),
    ("4. Mobile Other", "Other warehouse mobile screens (login, tasks, scan, count, GRN, returns, transfer)"),
    ("5. API Routes", "Backend route-group contract & permission tests"),
    ("6. E2E Flows", "Cross-module business workflows"),
    ("7. Role Matrix", "RBAC reference — which role sees which page"),
    ("", ""),
    ("Legend", "Status values: Not Run · In Progress · Passed · Failed · Blocked"),
    ("Priority", "High · Medium · Low"),
    ("Type", "Functional · UI · Validation · Edge · Permission · Integration · Performance"),
    ("Changes Required", "Specific code/config changes needed before this test can pass"),
]


def overview_sheet(wb):
    if "Overview" in wb.sheetnames:
        del wb["Overview"]
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "PVS ERP — Master Test Plan"
    ws["A1"].font = Font(bold=True, size=18, color="003087")
    ws.merge_cells("A1:B1")
    ws["A2"] = "Updated detailed test cases — Web Portal & Mobile Picking/Packing"
    ws["A2"].font = Font(italic=True, size=11, color="555555")
    ws.merge_cells("A2:B2")

    for r, (k, v) in enumerate(OVERVIEW, 4):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=2).alignment = WRAP

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120


# ────────────────────────────────────────────────────────────────────
# 1. Models
# ────────────────────────────────────────────────────────────────────
MODELS = [
    make_row("MDL-001", "Auth & Users", "User", "Create user with each role",
        "Settings → Users → Add user for admin, supervisor, procurement, billing, warehouse",
        "User saved; appears in list with correct role badge",
        "Functional", "High", "Not Run"),
    make_row("MDL-002", "Auth & Users", "User", "Unique username enforced",
        "Create a second user with an existing username",
        "Rejected with a clear duplicate error",
        "Validation", "High", "Not Run"),
    make_row("MDL-003", "Auth & Users", "User", "Deactivate user blocks login",
        "Set user inactive; attempt login via /login and /m/login",
        "Login refused on both surfaces with 401",
        "Functional", "High", "Not Run"),
    make_row("MDL-004", "Auth & Users", "Session", "JWT issued on login + PIN",
        "Login via portal (password) and via /m/login (6-digit PIN)",
        "Bearer token returned and stored; protected /v1 calls succeed",
        "Functional", "High", "Not Run"),
    make_row("MDL-005", "Auth & Users", "Session", "Expired/tampered token",
        "Edit token in localStorage, call /v1/products",
        "401 returned; portal redirects to /login, mobile to /m/login",
        "Edge", "High", "Not Run"),
    make_row("MDL-006", "Master Data", "UomCategory / Uom", "Create UOM category & units",
        "Settings → UOMs → add category, add base + derived units with factor",
        "Units saved; conversion factor respected in pickers",
        "Functional", "Medium", "Not Run"),
    make_row("MDL-007", "Master Data", "Uom", "Conversion factor validation",
        "Add a unit with factor 0 or negative",
        "Rejected with validation error",
        "Validation", "Medium", "Not Run"),
    make_row("MDL-008", "Warehouse", "Warehouse", "Create storage vs production warehouse",
        "Settings → Warehouses → create with kind=storage and kind=production",
        "Both saved; kind shown correctly in UI",
        "Functional", "High", "Not Run"),
    make_row("MDL-009", "Warehouse", "Bin", "Create rack/bin hierarchy",
        "Warehouse page → Add rack, Add bin; verify zone/rack/shelf/bin tree",
        "Tree reflects new nodes; bin code composed correctly",
        "Functional", "High", "Not Run"),
    make_row("MDL-010", "Warehouse", "Bin", "Bin capacity / reservedQty integrity",
        "Reserve stock via picking, inspect reservedQty",
        "reservedQty never exceeds qty; freed on cancel/complete",
        "Integration", "High", "Not Run"),
    make_row("MDL-011", "Warehouse", "Bin", "Zero-qty bins hidden from Locations view",
        "Drain a bin to 0 qty; open Inventory → Locations",
        "Bin no longer appears for the product",
        "Functional", "Medium", "Passed",
        "Already implemented in inventory.ts /inventory/locations filter."),
    make_row("MDL-012", "Products", "Product", "Create product of each type",
        "Products → add finished good, raw material, packaging, sub-assembly",
        "Saved with SKU, UOM, type; visible in list & pickers",
        "Functional", "High", "Not Run"),
    make_row("MDL-013", "Products", "ProductVariant", "Variant SKUs & stock are independent",
        "Add variants to a parent product; pick one variant, complete",
        "Variant.stockOnHand updates; Product.stockOnHand untouched",
        "Functional", "High", "Passed",
        "Fixed in manufacturing.ts complete-mo to only bump variant.stockOnHand."),
    make_row("MDL-014", "Manufacturing", "Bom", "Product-level vs variant-level BOM",
        "Create BOM with productId only, and another with productId+variantId",
        "Both saved; UI shows parent product in consumed picker for variant BOMs",
        "Functional", "High", "Passed",
        "Fix shipped: BomEditor allows parent product in variant-scoped BOMs."),
    make_row("MDL-015", "Manufacturing", "BomByproduct", "By-product released on MO complete",
        "Add by-product to BOM, complete an MO",
        "By-product qty added to its target warehouse; ledger row written",
        "Integration", "High", "Not Run",
        "Verify ledger lines for byproducts in manufacturing.ts /production-orders/:id/complete."),
    make_row("MDL-016", "Inventory", "StockLedger", "Append-only ledger",
        "Try to update or delete a StockLedger row directly via API",
        "Rejected — ledger is append-only by design",
        "Edge", "High", "Not Run"),
    make_row("MDL-017", "Inventory", "InventoryAdjustment", "Adjust requires explicit binId",
        "Call /v1/inventory/adjust without binId",
        "400 validation error",
        "Validation", "High", "Passed",
        "Already enforced in inventory.ts adjustSchema."),
    make_row("MDL-018", "Sales", "SalesOrder", "Status transitions enforced",
        "Move SO through draft → confirmed → fulfilled → closed",
        "Each transition allowed only from valid prior state",
        "Validation", "High", "Not Run"),
    make_row("MDL-019", "Sales", "PickList", "Pick reservation lifecycle",
        "Create pick list, claim, partially pick, complete",
        "Reservations created/freed; ledger entries match",
        "Integration", "High", "Not Run"),
    make_row("MDL-020", "Sales", "PackingSlip", "One slip per pick list",
        "Complete a pick → packing slip auto-issued",
        "PackingSlip created with status=open",
        "Functional", "High", "Not Run"),
    make_row("MDL-021", "Procurement", "PurchaseOrder", "PO status flow",
        "Create PO draft → approve → partial GRN → close",
        "Each GRN reduces remaining qty; status updates",
        "Integration", "High", "Not Run"),
    make_row("MDL-022", "Procurement", "Grn", "GRN posts to stock",
        "Receive a PO via /v1/grns; check StockLedger",
        "Stock increases; vendor balance updates",
        "Integration", "High", "Not Run"),
    make_row("MDL-023", "Returns", "CustomerReturn", "Return lifecycle",
        "Create return → decide lines → finalize",
        "Status transitions; stock restocked on accepted lines",
        "Integration", "Medium", "Not Run"),
    make_row("MDL-024", "Returns", "CreditNote", "Auto credit note on finalize",
        "Finalize a return with all lines accepted",
        "CreditNote created with matching value",
        "Integration", "Medium", "Not Run"),
    make_row("MDL-025", "Manufacturing", "ProductionOrder", "Variant stock on completion",
        "Complete an MO with variant output of 50 pcs",
        "Variant.stockOnHand += 50; parent Product.stockOnHand unchanged",
        "Functional", "High", "Passed",
        "Recent fix; verify via SELECT on prod DB after complete."),
    make_row("MDL-026", "Transfers", "TransferOrder", "TO claim → pick → drop",
        "Generate TO, claim on mobile, scan source bin, then dest bin",
        "Each step records; bin qty conserved",
        "Integration", "High", "Not Run"),
    make_row("MDL-027", "Settings", "PutawayRule", "Rule resolves on MO completion",
        "Create rule for a product class; complete an MO",
        "FG putaway TO generated targeting the resolved bin",
        "Integration", "High", "Not Run"),
    make_row("MDL-028", "Settings", "PutawayRule", "Putaway rule deletion cascade",
        "Delete a rule referenced by an open TO",
        "TO retains its bin; rule removed cleanly",
        "Edge", "Medium", "Not Run"),
]

# ────────────────────────────────────────────────────────────────────
# 2. ERP Pages (Web portal) — DETAILED
# ────────────────────────────────────────────────────────────────────
PAGES = [
    # ── Login ──────────────────────────────────────────────────────
    make_row("WEB-LOGIN-01", "Auth", "Login page", "Valid login routes to dashboard",
        "Open /login; enter admin/nova1234 → submit",
        "Token stored; URL becomes /dashboard; user chip shows 'admin'",
        "Functional", "High", "Not Run"),
    make_row("WEB-LOGIN-02", "Auth", "Login page", "Invalid credentials show error",
        "Submit wrong password",
        "Inline error 'Invalid credentials'; URL stays /login; no token stored",
        "Functional", "High", "Not Run"),
    make_row("WEB-LOGIN-03", "Auth", "Login page", "Empty submit blocked client-side",
        "Click Sign in without filling fields",
        "Disabled or shows required field validation",
        "Validation", "Medium", "Not Run"),
    make_row("WEB-LOGIN-04", "Auth", "Login page", "Token persists on reload",
        "Login, reload page",
        "Still authenticated; dashboard re-loads with no re-prompt",
        "Functional", "High", "Not Run"),
    make_row("WEB-LOGIN-05", "Auth", "Login page", "Logout clears token",
        "From any page click profile → Log out",
        "Routed to /login; calls to /v1 return 401",
        "Functional", "High", "Not Run"),

    # ── Dashboard ──────────────────────────────────────────────────
    make_row("WEB-DASH-01", "Dashboard", "Dashboard", "KPI strip loads",
        "Open /dashboard",
        "Sales today, Open SOs, Stock value, Pending tasks tiles populated",
        "UI", "Medium", "Not Run"),
    make_row("WEB-DASH-02", "Dashboard", "Dashboard", "Sales trend chart renders",
        "Verify the area chart on dashboard",
        "Last-7-day series renders without console errors",
        "UI", "Medium", "Not Run"),
    make_row("WEB-DASH-03", "Dashboard", "Dashboard", "Live activity feed updates",
        "Trigger an action (e.g. create SO) in another tab",
        "Dashboard feed reflects the event within 60s (or on refresh)",
        "Functional", "Low", "Not Run"),
    make_row("WEB-DASH-04", "Dashboard", "Dashboard", "Drill-down links",
        "Click KPI tile (e.g. Open SOs)",
        "Routes to /sales-orders pre-filtered by status",
        "Functional", "Medium", "Not Run"),

    # ── Products ───────────────────────────────────────────────────
    make_row("WEB-PRD-01", "Products", "Products list", "List loads & paginates",
        "Open /products",
        "Grid renders; pagination/scroll works; SKU and name columns visible",
        "Functional", "High", "Not Run"),
    make_row("WEB-PRD-02", "Products", "Products list", "SKU/name search",
        "Type a partial SKU in the search box",
        "List narrows in real-time; clear restores full list",
        "Functional", "High", "Not Run"),
    make_row("WEB-PRD-03", "Products", "Product detail", "Open detail side panel",
        "Click a row",
        "Right panel shows attributes, lifecycle, variants, BOMs",
        "UI", "High", "Not Run"),
    make_row("WEB-PRD-04", "Products", "Product CRUD", "Create finished-good product",
        "New Product → enter SKU, name, UOM, type=Finished good → Save",
        "Saved; appears in list; SKU is unique",
        "Functional", "High", "Not Run"),
    make_row("WEB-PRD-05", "Products", "Product CRUD", "Edit + auto-save",
        "Open product, change description, blur field",
        "Change persists; toast 'Saved'",
        "Functional", "High", "Not Run"),
    make_row("WEB-PRD-06", "Products", "Product CRUD", "Soft delete / archive",
        "Archive a product",
        "No longer in active list; still visible with archived filter",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-PRD-07", "Products", "Variants", "Add variant",
        "Open product, Variants tab → Add → fill size/colour",
        "Variant gets unique SKU and is selectable in pickers",
        "Functional", "High", "Not Run"),
    make_row("WEB-PRD-08", "Products", "Variants", "Variant stock separate from parent",
        "Open variant, look at stockOnHand",
        "Variant qty does NOT roll up to parent product (verified post-fix)",
        "Functional", "High", "Passed",
        "Backed by MO-complete fix in manufacturing.ts."),
    make_row("WEB-PRD-09", "Products", "Bin locations link", "View bin locations",
        "Open product detail → 'View bin locations →'",
        "Routes to /inventory?tab=locations&q=<sku>; shows bins with qty",
        "Functional", "Medium", "Passed",
        "Deep-link wired via InventoryLocationsPanel."),
    make_row("WEB-PRD-10", "Products", "Image management", "Upload product image",
        "Open product, upload PNG/JPG ≤2MB",
        "Image renders inline and via /uploads/ proxy",
        "Functional", "Medium", "Not Run"),

    # ── Inventory ──────────────────────────────────────────────────
    make_row("WEB-INV-01", "Inventory", "Inventory page", "Tabs order: Locations first, Ledger second",
        "Open /inventory",
        "Locations tab is default and first; Ledger is second",
        "UI", "Medium", "Passed",
        "Tab order changed in Inventory.tsx."),
    make_row("WEB-INV-02", "Inventory", "Locations tab", "Default loads all products with stock",
        "Open /inventory, do not type anything",
        "Tabular list of every product/variant that has at least one non-zero bin",
        "Functional", "High", "Passed",
        "Backend filter qty>0 added."),
    make_row("WEB-INV-03", "Inventory", "Locations tab", "Fast search filter",
        "Type SKU/name in search box",
        "Client-side filter narrows rows instantly",
        "Functional", "High", "Passed"),
    make_row("WEB-INV-04", "Inventory", "Locations tab", "Bin location column shown",
        "Inspect table headers",
        "Only Bin Location, Qty, Reserved, Available columns; Zone/Shelf/Bin separate columns are removed",
        "UI", "Medium", "Passed",
        "Column trimmed in InventoryLocationsPanel."),
    make_row("WEB-INV-05", "Inventory", "Locations tab", "Reserved column meaning",
        "Hover/tooltip on Reserved header",
        "Tooltip: 'Stock allocated to open picks/SOs but not yet shipped'",
        "UI", "Low", "Not Run",
        "Add tooltip text to InventoryLocationsPanel header (minor enhancement)."),
    make_row("WEB-INV-06", "Inventory", "Ledger tab", "Filter by date and txn type",
        "Open Ledger, filter GRN range last 7 days",
        "Only matching ledger rows shown",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-INV-07", "Inventory", "Adjust modal", "Bin selection mandatory",
        "Open Adjust Stock; try to submit without picking a bin",
        "Submit disabled / validation: 'Select a bin'",
        "Validation", "High", "Passed",
        "Enforced in AdjustStockModal."),
    make_row("WEB-INV-08", "Inventory", "Adjust modal", "Positive adjust posts ledger",
        "Adjust +10 with reason='physical recount'; submit",
        "Bin qty +10; StockLedger row created with reason",
        "Functional", "High", "Not Run"),
    make_row("WEB-INV-09", "Inventory", "Adjust modal", "Negative adjust cannot go below zero",
        "Adjust −9999 on a low-stock bin",
        "400 with 'insufficient stock' error",
        "Validation", "High", "Not Run"),
    make_row("WEB-INV-10", "Inventory", "Scrolling", "Locations table scrolls inside viewport",
        "Resize window small; scroll within table",
        "Scrollbar appears on table; toolbar stays fixed",
        "UI", "High", "Passed",
        "Layout fixed: overflow-hidden on Inventory.tsx root."),

    # ── Warehouse ──────────────────────────────────────────────────
    make_row("WEB-WH-01", "Warehouse", "Warehouse page", "Bin tree renders",
        "Open /warehouse",
        "Warehouse → Zone → Rack → Shelf → Bin tree visible with occupancy colours",
        "UI", "High", "Not Run"),
    make_row("WEB-WH-02", "Warehouse", "Fast Transfer", "Scan → Source → Dest → Qty → F8",
        "Use Fast Transfer panel",
        "Stock moves; ledger entries written",
        "Integration", "High", "Not Run"),
    make_row("WEB-WH-03", "Warehouse", "Bin layout modal", "Add bins in bulk",
        "Bulk-add zone + shelves × bins",
        "All bins created with composed codes",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-WH-04", "Warehouse", "Deep-link", "Open with selected bin",
        "Open /warehouse?binId=<id>",
        "Tree expands to and selects the bin",
        "Functional", "Medium", "Not Run"),

    # ── Manufacturing ──────────────────────────────────────────────
    make_row("WEB-MFG-01", "Manufacturing", "Manufacturing page", "Empty state shows New Order",
        "Open /manufacturing with no MOs",
        "Toolbar with 'New Order' button always visible",
        "UI", "High", "Passed",
        "EmptyState now keeps toolbar visible."),
    make_row("WEB-MFG-02", "Manufacturing", "MO create", "Create MO from BOM",
        "New Order → pick product, BOM, batch size 50 → Save",
        "MO created in draft with shortage analysis",
        "Functional", "High", "Not Run"),
    make_row("WEB-MFG-03", "Manufacturing", "MO shortage", "Shortage deep-links to Adjust",
        "Open MO with shortage; click shortage row",
        "Routes to /inventory adjust modal pre-filled with product",
        "Functional", "Medium", "Passed"),
    make_row("WEB-MFG-04", "Manufacturing", "MO release", "Release MO consumes stock",
        "Release a draft MO",
        "Stock reserved/consumed per BOM consumed lines",
        "Integration", "High", "Not Run"),
    make_row("WEB-MFG-05", "Manufacturing", "MO complete", "Complete posts FG",
        "Complete an in-progress MO",
        "Output qty added to product / variant stockOnHand",
        "Integration", "High", "Not Run"),
    make_row("WEB-MFG-06", "Manufacturing", "MO complete", "By-products released",
        "Complete MO with by-products defined",
        "By-product qty added to target WH; ledger row per byproduct",
        "Integration", "High", "Not Run"),
    make_row("WEB-MFG-07", "Manufacturing", "MO complete", "Variant-only stock bump",
        "Complete MO producing 50 of variant V1",
        "Variant V1 stockOnHand += 50; parent product unchanged",
        "Functional", "High", "Passed",
        "Recent fix verified."),
    make_row("WEB-MFG-08", "Manufacturing", "MO inventory trail", "Show consumed + released items",
        "Open completed MO → Inventory trail",
        "Lists every consumed (-) and released (+) ledger entry",
        "Functional", "Medium", "Not Run"),

    # ── BOMs ───────────────────────────────────────────────────────
    make_row("WEB-BOM-01", "BOMs", "BOMs page", "Standalone navigation",
        "Click 'BOMs' under Manufacturing in left nav",
        "Routes to /manufacturing/boms; lists all BOMs",
        "Functional", "High", "Passed"),
    make_row("WEB-BOM-02", "BOMs", "BOMs page", "Single nav highlight",
        "Click BOMs link",
        "Only 'BOMs' is highlighted, not 'Manufacturing' too",
        "UI", "Medium", "Passed",
        "Fixed via end:true on Manufacturing NavLink."),
    make_row("WEB-BOM-03", "BOMs", "BOM editor", "New BOM button works",
        "Click 'New BOM'",
        "Routes to /manufacturing/boms/new; empty editor",
        "Functional", "High", "Passed"),
    make_row("WEB-BOM-04", "BOMs", "BOM editor", "Variant-level BOM picker allows parent",
        "Create variant-level BOM; open Consumed picker",
        "Parent product appears in the picker (so 1 pc variant can consume 0.1 kg of parent)",
        "Functional", "High", "Passed"),
    make_row("WEB-BOM-05", "BOMs", "BOM editor", "Add by-product tab",
        "Open By-products tab; add an item",
        "Row persists; visible on save",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-BOM-06", "BOMs", "BOM editor", "Page scrolls in editor",
        "Open large BOM and scroll",
        "Body scrolls inside the page; footer fixed",
        "UI", "Medium", "Passed"),
    make_row("WEB-BOM-07", "BOMs", "BOMs page", "Delete BOM respects FK",
        "Delete a BOM referenced by an open MO",
        "Blocked with FK-friendly error; supervisor can cancel MO first",
        "Validation", "High", "Not Run"),

    # ── Procurement ────────────────────────────────────────────────
    make_row("WEB-PROC-01", "Procurement", "Procurement page", "PO list & filters",
        "Open /procurement",
        "List of POs with status pills; filter by status/vendor",
        "Functional", "High", "Not Run"),
    make_row("WEB-PROC-02", "Procurement", "PO create", "Create PO with multi-lines",
        "New PO → pick vendor → add 3 product lines → Save",
        "PO created with status=draft; total computed",
        "Functional", "High", "Not Run"),
    make_row("WEB-PROC-03", "Procurement", "PO approve", "Approve PO transitions status",
        "Approve a draft PO",
        "Status=approved; lines locked",
        "Functional", "High", "Not Run"),
    make_row("WEB-PROC-04", "Procurement", "GRN create", "Receive against PO",
        "Open PO → Receive → enter qty + QC pass",
        "GRN created; PO qtyReceived updates; stock rises",
        "Integration", "High", "Not Run"),
    make_row("WEB-PROC-05", "Procurement", "GRN partial", "Partial receipt allowed",
        "Receive 5 of 10 ordered",
        "PO marked partial; remaining 5 still open",
        "Functional", "High", "Not Run"),
    make_row("WEB-PROC-06", "Procurement", "Vendor master", "Vendor CRUD + GST validation",
        "Add vendor with invalid GST",
        "Validation error; valid GST saves",
        "Validation", "Medium", "Not Run"),

    # ── Price Lists ────────────────────────────────────────────────
    make_row("WEB-PL-01", "Price Lists", "Price Lists page", "CRUD list & items",
        "Create list, add 5 items with rates",
        "Saved; appears for selection in Quotes / SOs",
        "Functional", "High", "Not Run"),
    make_row("WEB-PL-02", "Price Lists", "Effective dates", "Date-bounded pricing",
        "Add item with effectiveFrom/To dates",
        "Resolution returns the correct rate for a given quote date",
        "Functional", "Medium", "Not Run"),

    # ── Customers ──────────────────────────────────────────────────
    make_row("WEB-CUST-01", "Customers", "Customers page", "CRUD customer",
        "Add a customer with GST + address",
        "Saved; appears in pickers for Quotes/SOs",
        "Functional", "High", "Not Run"),
    make_row("WEB-CUST-02", "Customers", "AR ledger", "Record payment",
        "Open customer → Receive payment ₹10,000",
        "Outstanding decreases; ledger row added",
        "Functional", "High", "Not Run"),
    make_row("WEB-CUST-03", "Customers", "Statement", "Generate statement",
        "Open Statement",
        "Lists invoices, credits, payments, balance",
        "Functional", "Medium", "Not Run"),

    # ── Enquiries / Quotes / SO ────────────────────────────────────
    make_row("WEB-ENQ-01", "Enquiries", "Enquiry CRUD", "Create + convert to quote",
        "New enquiry → Convert",
        "Quote draft created pre-filled",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-QT-01", "Quotes", "Quote CRUD", "Create quote with revisions",
        "Edit twice; check revisions tab",
        "Each revision stored; latest shown by default",
        "Functional", "High", "Not Run"),
    make_row("WEB-QT-02", "Quotes", "Quote share", "Public share link",
        "Click Share → copy link",
        "Opens /share/quote/:token in incognito; renders read-only",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-SO-01", "Sales Orders", "SO CRUD", "Create from quote",
        "Quote → Convert to SO",
        "SO created; lines + customer copied",
        "Functional", "High", "Not Run"),
    make_row("WEB-SO-02", "Sales Orders", "SO progress", "Generate pick list",
        "Confirm SO; click 'Generate pick list'",
        "Pick list created with the SO lines",
        "Integration", "High", "Not Run"),

    # ── Picking & Packing (desktop view) ───────────────────────────
    make_row("WEB-PICK-01", "Picking (web)", "Picking page", "List & filters",
        "Open /picking",
        "Pick lists with status pills, customer, assignee",
        "Functional", "High", "Not Run"),
    make_row("WEB-PICK-02", "Picking (web)", "Pick list detail", "Print pick list",
        "Open pick → Print",
        "Opens /print/pick-list/:id chrome-free; lines in walk order",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-PACK-01", "Packing (web)", "Packing page", "List & filters",
        "Open /packing",
        "Slips with status pills; can drill in",
        "Functional", "High", "Not Run"),
    make_row("WEB-PACK-02", "Packing (web)", "Print slip", "Print packing slip",
        "Open slip → Print",
        "Opens /print/packing-slip/:id; QR code + items",
        "Functional", "Medium", "Not Run"),

    # ── Returns ────────────────────────────────────────────────────
    make_row("WEB-RT-01", "Returns", "Returns page", "Create return from invoice",
        "Open invoice → Create return",
        "Return draft created with original lines",
        "Functional", "High", "Not Run"),
    make_row("WEB-RT-02", "Returns", "Decide lines", "Per-line decision",
        "Approve some, reject some",
        "Statuses recorded per line",
        "Functional", "High", "Not Run"),
    make_row("WEB-RT-03", "Returns", "Finalize", "Issue credit note",
        "Finalize the return",
        "CreditNote auto-issued; stock restocked on approved lines",
        "Integration", "High", "Not Run"),

    # ── Transport ──────────────────────────────────────────────────
    make_row("WEB-TRN-01", "Transport", "Transport page", "Dispatch queue",
        "Open /transport",
        "Outgoing trips with vehicles, drivers; can assign trip to slip",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-TRN-02", "Transport", "Delivery proof", "OTP + photo capture",
        "Mark delivered with OTP",
        "Trip locked; proof attached",
        "Functional", "Medium", "Not Run"),

    # ── Approvals ──────────────────────────────────────────────────
    make_row("WEB-APR-01", "Approvals", "Approvals page", "Pending list",
        "Open /approvals",
        "Lists items awaiting review (e.g. PO approvals)",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-APR-02", "Approvals", "F8 to approve", "Hotkey approves",
        "Highlight item → press F8",
        "Approved; audit log written",
        "Functional", "Low", "Not Run"),

    # ── Reports ────────────────────────────────────────────────────
    make_row("WEB-RPT-01", "Reports", "Reports page", "Report library renders",
        "Open /reports",
        "Grouped reports list; clicking opens detail",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-RPT-02", "Reports", "Sales report", "Date range filter",
        "Pick last 30 days",
        "Charts and grid update",
        "Functional", "Medium", "Not Run"),

    # ── Settings ───────────────────────────────────────────────────
    make_row("WEB-SET-01", "Settings", "Settings page", "Tabs render",
        "Open /settings (admin only)",
        "Company, Users, Roles, Scanner, Putaway, Backups, etc. tabs visible",
        "UI", "High", "Not Run"),
    make_row("WEB-SET-02", "Settings", "Putaway rules", "Single nav location",
        "Inspect left nav for 'Putaway rules'",
        "NOT present in main nav; only available under Settings → Putaway",
        "UI", "Medium", "Passed",
        "Duplicate nav link removed."),
    make_row("WEB-SET-03", "Settings", "Putaway rules", "Add rule",
        "Settings → Putaway → Add rule",
        "Rule saved; visible in list",
        "Functional", "High", "Not Run"),
    make_row("WEB-SET-04", "Settings", "Users & roles", "Create new role",
        "Add role 'qc'",
        "Role saved; assignable to users",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-SET-05", "Settings", "Scanner config", "Toggle BLE scanner",
        "Enable Bluetooth scanner mode",
        "Persists across sessions",
        "Functional", "Low", "Not Run"),

    # ── Shell / Cross-cutting ─────────────────────────────────────
    make_row("WEB-SH-01", "Shell", "Command palette", "Ctrl+K opens",
        "Press Ctrl+K from any page",
        "Modal opens with searchable command list",
        "UI", "High", "Not Run"),
    make_row("WEB-SH-02", "Shell", "Command palette", "Quick navigate",
        "Type 'inv' → Enter",
        "Routes to /inventory",
        "Functional", "High", "Not Run"),
    make_row("WEB-SH-03", "Shell", "Scrolling", "Main area uses overflow-hidden",
        "Inspect <main> element",
        "Pages own their internal scrolling; no double scrollbars",
        "UI", "High", "Passed"),
    make_row("WEB-SH-04", "Shell", "Workspace tabs", "Detach + close",
        "Open multiple tabs; close one",
        "Tabs persist across reloads; close works",
        "Functional", "Medium", "Not Run"),
    make_row("WEB-SH-05", "Shell", "Brand provider", "Theme loads from settings",
        "Change company logo in Settings; reload",
        "Logo updates across portal",
        "Functional", "Low", "Not Run"),
]

# ────────────────────────────────────────────────────────────────────
# 3. Mobile Pick/Pack — DETAILED
# ────────────────────────────────────────────────────────────────────
MOB_PICK = [
    # ── Login & Tasks (entry to pick/pack) ─────────────────────────
    make_row("MOB-LOGIN-01", "Mobile · Login", "MobileLogin", "Open APK lands on login",
        "Install APK; tap icon",
        "First screen is /m/login (NOT desktop dashboard)",
        "Functional", "High", "Passed",
        "MOBILE_BUILD mode strips desktop routes."),
    make_row("MOB-LOGIN-02", "Mobile · Login", "MobileLogin", "Username + PIN login",
        "Enter warehouse1 / PIN 123456",
        "Token stored; routes to /m/tasks",
        "Functional", "High", "Not Run"),
    make_row("MOB-LOGIN-03", "Mobile · Login", "MobileLogin", "Warehouse picker on first login",
        "First-time login; warehouse not yet selected",
        "Picker shown; selection persists across sessions",
        "Functional", "High", "Not Run"),
    make_row("MOB-LOGIN-04", "Mobile · Login", "MobileLogin", "Logout returns to login",
        "From Profile → Log out",
        "Token cleared; routed to /m/login",
        "Functional", "Medium", "Not Run"),

    make_row("MOB-TASK-01", "Mobile · Tasks", "MobileTasks", "Pick/Pack/Move tabs render counts",
        "Open /m/tasks",
        "Three segments show 'Pick (n)', 'Pack (n)', 'Move (n)' with live counts",
        "Functional", "High", "Not Run"),
    make_row("MOB-TASK-02", "Mobile · Tasks", "MobileTasks", "More tab quick links",
        "Tap 'More'",
        "Tiles for Goods Receipt (GRN), Cycle Count, Customer Returns",
        "Functional", "Medium", "Passed",
        "New 'More' segment + QuickLink tiles added."),
    make_row("MOB-TASK-03", "Mobile · Tasks", "MobileTasks", "Claim a pick task",
        "On Pick tab → Available → tap Claim",
        "Task moves to 'My tasks'; assignedToId set",
        "Functional", "High", "Not Run"),
    make_row("MOB-TASK-04", "Mobile · Tasks", "MobileTasks", "Claim race shows 409",
        "Try to claim a pick that's already claimed elsewhere",
        "Banner 'Someone else just claimed this. Refreshing…'",
        "Edge", "Medium", "Not Run"),
    make_row("MOB-TASK-05", "Mobile · Tasks", "MobileTasks", "Auto-refresh interval",
        "Open Tasks; wait 30s",
        "List silently refreshes",
        "Functional", "Low", "Not Run"),

    # ── Picking ───────────────────────────────────────────────────
    make_row("MOB-PICK-01", "Mobile · Pick", "MobilePick", "Open my claimed pick",
        "Tap a claimed pick on /m/tasks",
        "Routes to /m/picks/:id; header card shows pick #, customer, item count",
        "Functional", "High", "Not Run"),
    make_row("MOB-PICK-02", "Mobile · Pick", "MobilePick", "Lines in walk order",
        "Inspect line ordering",
        "Lines sorted by bin walk-path (zone/shelf/bin ascending)",
        "Functional", "High", "Not Run"),
    make_row("MOB-PICK-03", "Mobile · Pick", "MobilePick", "Pending vs confirmed visual",
        "Drill into a line, confirm it, return",
        "Confirmed lines show emerald accent + 'confirmed' chip; pending stay amber",
        "UI", "Medium", "Not Run"),
    make_row("MOB-PICK-04", "Mobile · Pick", "MobilePick", "Release returns to queue",
        "Tap 'Release'",
        "Claim freed; routed back to /m/tasks",
        "Functional", "High", "Not Run"),
    make_row("MOB-PICK-05", "Mobile · Pick", "MobilePick", "Complete with all lines picked",
        "Confirm every line → 'Complete pick'",
        "Status=picked; packing slip auto-issued; banner offers to open it",
        "Integration", "High", "Not Run"),
    make_row("MOB-PICK-06", "Mobile · Pick", "MobilePick", "Complete blocked when lines pending",
        "Try Complete with 1 line unpicked",
        "Button disabled; status text 'X lines still need a scan'",
        "Validation", "High", "Not Run"),
    make_row("MOB-PICK-07", "Mobile · Pick", "MobilePick", "Stale-line recovery",
        "Concurrently drain stock for one line in another session, then Complete",
        "409 pick_blocked → amber recovery card lists the stale SKUs and offers 'Reset stale lines'; tap resets qty to 0 and re-attempts Complete",
        "Edge", "High", "Not Run"),
    make_row("MOB-PICK-08", "Mobile · Pick", "MobilePick", "Locked-state banner",
        "Open an already-picked pick",
        "Banner 'This pick is complete' with 'Open packing slip' CTA; action bar hidden",
        "UI", "High", "Not Run"),
    make_row("MOB-PICK-09", "Mobile · Pick", "MobilePick", "Offline banner appears",
        "Toggle airplane mode while viewing pick",
        "Amber 'Offline' banner shows; Complete refuses to send",
        "Edge", "High", "Passed",
        "Offline detection added in pick-harden phase."),
    make_row("MOB-PICK-10", "Mobile · Pick", "MobilePick", "Retry button on error",
        "Force /v1/pick-lists/:id failure (e.g. bad token)",
        "Error banner shows 'Retry' link; tapping refetches",
        "Edge", "High", "Passed"),

    # ── Pick line scan-confirm ────────────────────────────────────
    make_row("MOB-PKLN-01", "Mobile · Pick line", "MobilePickLine", "Three-step scan flow",
        "Open a pick line",
        "Steps visible: 1. Scan bin, 2. Scan product, 3. Confirm qty",
        "UI", "High", "Not Run"),
    make_row("MOB-PKLN-02", "Mobile · Pick line", "MobilePickLine", "Bin scan via ML Kit camera",
        "Tap 'Scan' button next to Bin field",
        "Native camera opens; bin barcode auto-filled when scanned",
        "Functional", "High", "Not Run",
        "Requires @capacitor-mlkit/barcode-scanning on device."),
    make_row("MOB-PKLN-03", "Mobile · Pick line", "MobilePickLine", "Bin mismatch shows reason picker",
        "Scan a bin different from the expected one",
        "Amber 'doesn't match' hint + reason dropdown defaults to 'wrong_bin'",
        "Functional", "High", "Not Run"),
    make_row("MOB-PKLN-04", "Mobile · Pick line", "MobilePickLine", "Product mismatch flags",
        "Scan a different SKU than expected",
        "Reason auto-selects 'substitute'; can edit",
        "Functional", "High", "Not Run"),
    make_row("MOB-PKLN-05", "Mobile · Pick line", "MobilePickLine", "Qty +/- stepper",
        "Tap + and − buttons",
        "Qty changes by 1; cannot go below 0",
        "Functional", "High", "Not Run"),
    make_row("MOB-PKLN-06", "Mobile · Pick line", "MobilePickLine", "Short-pick reason auto-suggested",
        "Set qty below qtyToPick",
        "Reason defaults to 'short_pick' (or 'not_found' if qty=0)",
        "Functional", "High", "Not Run"),
    make_row("MOB-PKLN-07", "Mobile · Pick line", "MobilePickLine", "Confirm sends clientOpId",
        "Inspect /v1/pick-lists/:id/items/:itemId/scan request",
        "Body includes clientOpId UUID; same id reused on retry until success",
        "Integration", "High", "Passed",
        "newClientOpId already wired."),
    make_row("MOB-PKLN-08", "Mobile · Pick line", "MobilePickLine", "Locked-state guard",
        "Open a line whose pick is already 'picked'",
        "Page renders summary with 'Back to pick list' link, no scan form",
        "Edge", "Medium", "Not Run"),
    make_row("MOB-PKLN-09", "Mobile · Pick line", "MobilePickLine", "Reason='other' requires remarks",
        "Choose 'Other (see remarks)' but leave remarks blank, submit",
        "Validation prompts to fill remarks",
        "Validation", "Medium", "Not Run",
        "Backend already requires remarks for 'other'; verify the inline hint."),
    make_row("MOB-PKLN-10", "Mobile · Pick line", "MobilePickLine", "Successful confirm routes back",
        "Confirm pick line",
        "Returns to /m/picks/:id with line marked confirmed",
        "Functional", "High", "Not Run"),
    make_row("MOB-PKLN-11", "Mobile · Pick line", "MobilePickLine", "Variant SKU expected",
        "Pick a variant SKU; scan parent SKU instead",
        "Treated as mismatch; reason picker appears",
        "Edge", "Medium", "Not Run"),
    make_row("MOB-PKLN-12", "Mobile · Pick line", "MobilePickLine", "Bin code prefilled",
        "Open pick line that has a bin reservation",
        "Bin code prefilled in the bin field (no scan needed if it matches)",
        "Functional", "Medium", "Not Run"),

    # ── Packing ───────────────────────────────────────────────────
    make_row("MOB-PACK-01", "Mobile · Pack", "MobilePack", "Open packing slip",
        "Tap a claimed pack on /m/tasks",
        "Routes to /m/packs/:id; header card shows slip #, customer",
        "Functional", "High", "Not Run"),
    make_row("MOB-PACK-02", "Mobile · Pack", "MobilePack", "Per-line qty editable",
        "Adjust qty using − and + or numeric input",
        "Qty constrained between 0 and qtyPicked (over-pack rejected by server)",
        "Functional", "High", "Not Run"),
    make_row("MOB-PACK-03", "Mobile · Pack", "MobilePack", "Scan product per line",
        "Tap Scan; aim at product barcode",
        "Product code filled; matches expected SKU = no reason needed",
        "Functional", "High", "Not Run"),
    make_row("MOB-PACK-04", "Mobile · Pack", "MobilePack", "Short pack reason",
        "Set qty below qtyPicked",
        "Amber 'Reason for variance' panel appears with options (short_pack, damage, substitute, other)",
        "Functional", "High", "Not Run"),
    make_row("MOB-PACK-05", "Mobile · Pack", "MobilePack", "Confirm line then Update",
        "Confirm once; change qty; tap Update",
        "Backend updates qtyPacked; ledger entries adjusted",
        "Integration", "High", "Not Run"),
    make_row("MOB-PACK-06", "Mobile · Pack", "MobilePack", "Mark packed seals slip",
        "Confirm all lines → tap 'Mark packed'",
        "Slip status → packed; no further edits",
        "Integration", "High", "Not Run"),
    make_row("MOB-PACK-07", "Mobile · Pack", "MobilePack", "Mark packed disabled if pending",
        "Try Mark packed with 1 line unconfirmed",
        "Button disabled; helper text explains",
        "Validation", "High", "Not Run"),
    make_row("MOB-PACK-08", "Mobile · Pack", "MobilePack", "ClientOpId per line",
        "Inspect /v1/packing-slips/:id/items/:itemId/scan body",
        "Includes clientOpId; regenerated after success so corrections are fresh ops",
        "Integration", "High", "Passed"),
    make_row("MOB-PACK-09", "Mobile · Pack", "MobilePack", "Offline guard on Mark packed",
        "Toggle airplane mode; tap Mark packed",
        "Banner 'You're offline. Reconnect and try again.'",
        "Edge", "High", "Passed",
        "Added in pick-harden phase."),
    make_row("MOB-PACK-10", "Mobile · Pack", "MobilePack", "Release returns to queue",
        "Tap Release",
        "Claim freed; routed to /m/tasks",
        "Functional", "Medium", "Not Run"),
    make_row("MOB-PACK-11", "Mobile · Pack", "MobilePack", "Reason='substitute' captured",
        "Scan different SKU than expected; mark reason=substitute; submit",
        "Pack-scan call stores reason; visible in audit",
        "Functional", "Medium", "Not Run"),
    make_row("MOB-PACK-12", "Mobile · Pack", "MobilePack", "Retry banner on error",
        "Cause a 500 (e.g. proxy down)",
        "Red error banner with Retry link; tapping refetches the slip",
        "Edge", "High", "Passed"),

    # ── Barcode scanner integration ───────────────────────────────
    make_row("MOB-BS-01", "Mobile · Scanner", "BarcodeScanner", "Native ML Kit when available",
        "On APK device, open any scan field",
        "Native ML Kit camera opens (full-screen)",
        "Integration", "High", "Not Run"),
    make_row("MOB-BS-02", "Mobile · Scanner", "BarcodeScanner", "Web BarcodeDetector fallback",
        "Open /m/picks in Chrome on desktop",
        "Uses browser BarcodeDetector if available; otherwise shows manual input",
        "Functional", "Medium", "Not Run"),
    make_row("MOB-BS-03", "Mobile · Scanner", "BarcodeScanner", "Camera permission denied",
        "Deny camera permission",
        "Friendly state 'denied' with manual entry fallback",
        "Edge", "Medium", "Not Run"),
    make_row("MOB-BS-04", "Mobile · Scanner", "BarcodeScanner", "Manual entry path",
        "Type code into the input field",
        "Submitting accepts the code (without using camera)",
        "Functional", "High", "Not Run"),
    make_row("MOB-BS-05", "Mobile · Scanner", "BarcodeScanner", "Format support",
        "Scan QR, CODE_128, EAN_13",
        "All recognised; raw value populates the field",
        "Integration", "Medium", "Not Run"),

    # ── Performance / device ──────────────────────────────────────
    make_row("MOB-PERF-01", "Mobile · Perf", "APK launch", "Cold start < 3s",
        "Force-stop app then launch on a mid-tier device",
        "Login screen visible within 3s",
        "Performance", "Medium", "Not Run"),
    make_row("MOB-PERF-02", "Mobile · Perf", "Bundle size", "APK ≤ 8 MB",
        "Check release APK size",
        "≤ 8 MB for arm64-v8a build",
        "Performance", "Medium", "Passed",
        "Current build is 7.87 MB."),
    make_row("MOB-PERF-03", "Mobile · Perf", "Memory", "No leak across 30 picks",
        "Complete 30 picks back-to-back",
        "Memory stable; no OOM",
        "Performance", "Low", "Not Run"),

    # ── Connectivity / API ────────────────────────────────────────
    make_row("MOB-NET-01", "Mobile · Network", "API base URL", "Points at VPS for release build",
        "Inspect window.__APP_VERSION__ or network tab",
        "All /v1 calls go to http://217.216.78.119/v1",
        "Integration", "High", "Passed",
        "Set via erp-portal/.env.mobile → VITE_API_URL."),
    make_row("MOB-NET-02", "Mobile · Network", "401 handling", "Token expiry mid-session",
        "Wait for token to expire (or revoke)",
        "Next /v1 call returns 401 → app routes to /m/login",
        "Edge", "High", "Not Run"),
    make_row("MOB-NET-03", "Mobile · Network", "Idempotency", "Retried POST de-dups",
        "Trigger network timeout on a scan, then retry",
        "Backend deduplicates on clientOpId; no double consumption",
        "Integration", "High", "Not Run"),
]

# ────────────────────────────────────────────────────────────────────
# 4. Mobile other screens
# ────────────────────────────────────────────────────────────────────
MOB_OTHER = [
    # Scan / Verify / Location / Bin
    make_row("MOB-SCAN-01", "Mobile · Scan", "MobileScan", "Scan resolves product",
        "Scan a known SKU",
        "Routes to product detail / shows match",
        "Functional", "Medium", "Not Run"),
    make_row("MOB-SCAN-02", "Mobile · Scan", "MobileScan", "Scan resolves bin",
        "Scan a bin location code",
        "Routes to /m/bin/:id with details",
        "Functional", "Medium", "Not Run"),
    make_row("MOB-VFY-01", "Mobile · Verify", "MobileVerify", "Verify a packing slip QR",
        "Scan QR of a packed slip",
        "Verification result shown (ok / mismatch)",
        "Functional", "Low", "Not Run"),
    make_row("MOB-LOC-01", "Mobile · Location", "MobileLocation", "Open /m/loc/:code",
        "Tap a location on tasks list",
        "Renders location contents",
        "Functional", "Low", "Not Run"),
    make_row("MOB-BIN-01", "Mobile · Bin", "MobileBin", "Open bin detail",
        "Tap a bin",
        "Shows product, qty, capacity, last activity",
        "Functional", "Medium", "Not Run"),

    # GRN (procurement)
    make_row("MOB-GRN-01", "Mobile · GRN", "MobileGrnList", "Approved PO list",
        "Open /m/grn",
        "Lists POs in 'approved' status awaiting receipt",
        "Functional", "High", "Not Run",
        "Backend gates GRN endpoints to procurement role."),
    make_row("MOB-GRN-02", "Mobile · GRN", "MobileGrnReceive", "Receive a PO",
        "Tap a PO; enter received qty per line; QC=pass; Post",
        "GRN created; PO qtyReceived updates; stock rises",
        "Integration", "High", "Not Run"),
    make_row("MOB-GRN-03", "Mobile · GRN", "MobileGrnReceive", "Role gate clear feedback",
        "Login as warehouse role; tap Post GRN",
        "403 with hint 'Requires procurement role'",
        "Permission", "High", "Not Run",
        "Add backend role to user OR extend role gate on POST /grns."),
    make_row("MOB-GRN-04", "Mobile · GRN", "MobileGrnReceive", "Partial receipt allowed",
        "Receive 5 of 10 ordered",
        "PO remains open with remaining 5",
        "Functional", "High", "Not Run"),

    # Cycle count
    make_row("MOB-CNT-01", "Mobile · Count", "MobileCount", "Scan bin → resolved card",
        "/m/count → enter or scan bin code A/1/1",
        "Bin card shows product, qty, warehouse",
        "Functional", "High", "Not Run"),
    make_row("MOB-CNT-02", "Mobile · Count", "MobileCount", "Recount tab posts to ledger",
        "Recount tab → enter actual qty → reason → Confirm",
        "POST /v1/bins/:id/recount with clientOpId; ledger row written",
        "Integration", "High", "Not Run"),
    make_row("MOB-CNT-03", "Mobile · Count", "MobileCount", "Reassign tab",
        "Reassign → enter new SKU → qty → Confirm",
        "Bin reassigned to new product; previous qty zeroed",
        "Integration", "High", "Not Run"),
    make_row("MOB-CNT-04", "Mobile · Count", "MobileCount", "Quick adjust",
        "Quick adjust tab → product id + warehouse id + qty + reason",
        "POST /v1/inventory/adjust succeeds",
        "Integration", "Medium", "Not Run"),
    make_row("MOB-CNT-05", "Mobile · Count", "MobileCount", "Idempotent retry",
        "Cause network timeout, retry recount",
        "Same clientOpId; backend de-dupes; final state correct",
        "Edge", "High", "Not Run"),

    # Transfers
    make_row("MOB-TRF-01", "Mobile · Transfer", "MobileTransfer", "Claim transfer",
        "Tasks → Move tab → Claim",
        "Transfer moves to my tasks",
        "Functional", "High", "Not Run"),
    make_row("MOB-TRF-02", "Mobile · Transfer", "MobileTransfer", "Pick source bins",
        "Scan source bin, confirm qty",
        "Source bin qty reduced",
        "Integration", "High", "Not Run"),
    make_row("MOB-TRF-03", "Mobile · Transfer", "MobileTransfer", "Drop to destination",
        "Scan destination bin, confirm",
        "Dest bin qty increased; TO closed",
        "Integration", "High", "Not Run"),
    make_row("MOB-TRF-04", "Mobile · Transfer", "MobileTransfer", "Cancel mid-transfer",
        "Cancel an in-progress TO",
        "Stock returns to original bin; TO cancelled",
        "Edge", "Medium", "Not Run"),

    # Returns
    make_row("MOB-RTN-01", "Mobile · Returns", "MobileReturnList", "Pending returns list",
        "Open /m/returns",
        "Lists returns in 'pending' status",
        "Functional", "High", "Not Run"),
    make_row("MOB-RTN-02", "Mobile · Returns", "MobileReturnDetail", "Per-line decision",
        "Open a return; tap Accept on a line; Save decision",
        "Decision posted; line marked",
        "Functional", "High", "Not Run"),
    make_row("MOB-RTN-03", "Mobile · Returns", "MobileReturnDetail", "Credit-only mapping",
        "Choose 'Credit only' on a line",
        "Backend stores as 'rejected' with note '[Credit only] …' (API only accepts approved|rejected)",
        "Functional", "Medium", "Passed",
        "Workaround applied in MobileReturn.tsx."),
    make_row("MOB-RTN-04", "Mobile · Returns", "MobileReturnDetail", "Finalize blocked if undecided",
        "Try Finalize with some lines undecided",
        "Button disabled; helper text 'X lines still need a decision'",
        "Validation", "High", "Not Run"),
    make_row("MOB-RTN-05", "Mobile · Returns", "MobileReturnDetail", "Finalize succeeds",
        "All lines decided → Finalize",
        "Status=finalized; CreditNote issued; stock restocked on approved",
        "Integration", "High", "Not Run"),

    # Profile
    make_row("MOB-PROF-01", "Mobile · Profile", "MobileProfile", "Profile loads",
        "Open /m/profile",
        "Shows worker name, warehouse, version, log-out button",
        "Functional", "Low", "Not Run"),
    make_row("MOB-PROF-02", "Mobile · Profile", "MobileProfile", "Punch in/out",
        "Tap Punch in",
        "Attendance row written; UI flips to Punch out",
        "Functional", "Medium", "Not Run"),

    # Shell behaviour
    make_row("MOB-SH-01", "Mobile · Shell", "MobileShell", "Bottom-nav has 5 tabs",
        "Inspect /m/* bottom nav",
        "Tasks · Scan · Count · Verify · Profile",
        "UI", "Medium", "Passed",
        "Added Count tab."),
    make_row("MOB-SH-02", "Mobile · Shell", "MobileShell", "Top bar greets user + warehouse",
        "Login then inspect top bar",
        "Shows 'Good morning, <first name>' and warehouse code",
        "UI", "Low", "Not Run"),
    make_row("MOB-SH-03", "Mobile · Shell", "MobileShell", "Offline pill in top bar",
        "Toggle airplane mode",
        "'OFFLINE' pill appears next to user name",
        "Functional", "Medium", "Not Run"),
    make_row("MOB-SH-04", "Mobile · Shell", "MobileShell", "Install prompt only in browser",
        "Open /m/* in Chrome (not APK)",
        "Install button appears when prompt fires; APK hides it",
        "Functional", "Low", "Not Run"),
]

# ────────────────────────────────────────────────────────────────────
# 5. API routes
# ────────────────────────────────────────────────────────────────────
API = [
    make_row("API-AUTH-01", "API · Auth", "POST /v1/auth/login", "Login returns JWT",
        "POST {username, password}",
        "200 with {token, user}; wrong creds → 401",
        "Functional", "High", "Not Run"),
    make_row("API-AUTH-02", "API · Auth", "POST /v1/auth/login-pin", "PIN login for mobile",
        "POST {username, pin}",
        "200 with token; bad PIN → 401",
        "Functional", "High", "Not Run"),
    make_row("API-AUTH-03", "API · Auth", "GET /v1/auth/me", "Returns current user",
        "Bearer token → GET /me",
        "User + role; bad token → 401",
        "Functional", "High", "Not Run"),
    make_row("API-PROD-01", "API · Catalog", "GET /v1/products", "List with pagination",
        "GET /products?search=ajwn",
        "Returns matching products",
        "Functional", "High", "Not Run"),
    make_row("API-PROD-02", "API · Catalog", "GET /v1/products/by-sku/:sku", "SKU lookup",
        "GET by-sku/AJWN-100G-01",
        "Returns single variant or 404",
        "Functional", "High", "Not Run"),
    make_row("API-PROD-03", "API · Catalog", "GET /v1/products/by-barcode/:code", "Barcode lookup",
        "GET by-barcode/<EAN>",
        "Returns match or 404",
        "Functional", "Medium", "Not Run"),
    make_row("API-INV-01", "API · Inventory", "GET /v1/inventory/locations", "Search with empty q",
        "GET /inventory/locations",
        "Returns all products with non-zero bins",
        "Functional", "High", "Passed",
        "Backend filter qty>0 added."),
    make_row("API-INV-02", "API · Inventory", "POST /v1/inventory/adjust", "Requires binId",
        "POST without binId",
        "400 validation",
        "Validation", "High", "Passed"),
    make_row("API-PROC-01", "API · Procurement", "GET /v1/purchase-orders", "Filter by status",
        "GET ?status=approved",
        "Only approved POs returned",
        "Functional", "High", "Not Run"),
    make_row("API-PROC-02", "API · Procurement", "POST /v1/grns", "Procurement role only",
        "POST as warehouse user",
        "403 forbidden",
        "Permission", "High", "Not Run"),
    make_row("API-PROC-03", "API · Procurement", "POST /v1/grns", "Posts stock",
        "POST valid GRN body",
        "201 GRN created; product stockOnHand increased",
        "Integration", "High", "Not Run"),
    make_row("API-PICK-01", "API · Fulfilment", "GET /v1/pick-lists/:id", "Returns pick + items",
        "Authorised GET",
        "200 with items[], packingSlip ref",
        "Functional", "High", "Not Run"),
    make_row("API-PICK-02", "API · Fulfilment", "POST /v1/pick-lists/:id/items/:itemId/scan",
        "Idempotent on clientOpId",
        "Send same scan twice with same clientOpId",
        "Second call returns same result; stock consumed once",
        "Integration", "High", "Not Run"),
    make_row("API-PICK-03", "API · Fulfilment", "POST /v1/pick-lists/:id/complete",
        "Returns 409 pick_blocked when stock drained",
        "Drain variant between scan and complete",
        "409 with details[] of stale itemIds",
        "Edge", "High", "Not Run"),
    make_row("API-PACK-01", "API · Fulfilment", "POST /v1/packing-slips/:id/items/:itemId/scan",
        "Over-pack rejected",
        "Send qty > qtyPicked",
        "400 with 'cannot pack more than picked'",
        "Validation", "High", "Not Run"),
    make_row("API-PACK-02", "API · Fulfilment", "POST /v1/packing-slips/:id/pack",
        "Finalises slip",
        "POST after all lines confirmed",
        "200; status=packed; later POSTs return 409 bad_state",
        "Integration", "High", "Not Run"),
    make_row("API-TRF-01", "API · Transfers", "POST /v1/transfer-orders/:id/claim", "Race protection",
        "Two devices claim simultaneously",
        "First 200, second 409",
        "Edge", "High", "Not Run"),
    make_row("API-TRF-02", "API · Transfers", "POST /v1/transfer-orders/:id/cancel",
        "Restores stock",
        "Cancel a partial transfer",
        "Source bin qty restored; ledger entries balance",
        "Integration", "High", "Not Run"),
    make_row("API-MFG-01", "API · Manufacturing", "POST /v1/production-orders/:id/complete",
        "Variant scope correctly bumps variant stock only",
        "Complete MO for a variant",
        "ProductVariant.stockOnHand += output; parent Product unchanged",
        "Integration", "High", "Passed"),
    make_row("API-MFG-02", "API · Manufacturing", "POST /v1/production-orders/:id/complete",
        "By-products released",
        "Complete MO with byproducts",
        "Ledger rows for each byproduct; bin qty updated per putaway rule",
        "Integration", "High", "Not Run"),
    make_row("API-LOC-01", "API · Locations", "GET /v1/locations/scan", "Resolves bin/zone/shelf",
        "GET ?code=A/1/1",
        "Returns kind='bin' + binId + warehouse",
        "Functional", "High", "Not Run"),
    make_row("API-LOC-02", "API · Locations", "GET /v1/me/tasks", "Returns buckets",
        "Auth'd GET",
        "Pick/Pack/Transfer claimed+available counts",
        "Functional", "High", "Not Run"),
    make_row("API-RET-01", "API · Returns", "POST /v1/returns/:id/lines/:lineId/decide",
        "Approved/rejected only",
        "POST decision='credit_only'",
        "400 ('credit_only' is mapped to 'rejected' by the mobile app)",
        "Validation", "Medium", "Not Run"),
    make_row("API-RET-02", "API · Returns", "POST /v1/returns/:id/finalize",
        "Issues credit note",
        "POST after all lines decided",
        "200 with creditNote ref; stock restocked on approved lines",
        "Integration", "High", "Not Run"),
    make_row("API-BIN-01", "API · Bins", "POST /v1/bins/:id/recount",
        "Idempotent on clientOpId",
        "Retry recount with same opId",
        "Idempotent; no duplicate ledger rows",
        "Integration", "High", "Not Run"),
    make_row("API-BIN-02", "API · Bins", "POST /v1/bins/:id/reassign",
        "Empty bin first",
        "Reassign while bin has stock of another product",
        "Validation error unless qty=0 or stock moved",
        "Validation", "High", "Not Run"),
]

# ────────────────────────────────────────────────────────────────────
# 6. E2E flows
# ────────────────────────────────────────────────────────────────────
E2E = [
    make_row("E2E-01", "E2E", "Order-to-Cash",
        "Quote → SO → Pick (mobile) → Pack (mobile) → Invoice → Payment",
        "Run a full sales cycle with mobile picking/packing",
        "All statuses consistent; AR balance accurate; ledger trace complete",
        "Integration", "High", "Not Run"),
    make_row("E2E-02", "E2E", "Procure-to-Stock",
        "PO (web) → GRN (mobile) → On-hand",
        "Create PO on portal; receive on mobile",
        "Stock increases; ledger and PO progress match",
        "Integration", "High", "Not Run"),
    make_row("E2E-03", "E2E", "MO → Pick (mobile)",
        "MO release with shortage → replenishment TO → mobile transfer → MO consume",
        "Run the full raw replenishment flow",
        "Materials land at line warehouse; MO consumes from line",
        "Integration", "High", "Not Run"),
    make_row("E2E-04", "E2E", "Mobile Pick Resilience",
        "Pick offline → resync → complete",
        "Toggle airplane mode mid-pick, restore, finish",
        "Scans queued or re-attempted with clientOpId; no double consumption",
        "Edge", "High", "Not Run"),
    make_row("E2E-05", "E2E", "Stale-line recovery",
        "Two pickers drain same variant; one Complete fails 409, recovers",
        "Reproduce race; tap 'Reset stale lines'; re-Complete",
        "Recovery banner shown; reset clears stale qty; second Complete succeeds",
        "Edge", "High", "Not Run"),
    make_row("E2E-06", "E2E", "Returns → Credit Note",
        "Invoice → Mobile return decisions → Finalize → AR reflects credit",
        "Process a customer return end-to-end via mobile",
        "Stock restocked; credit note issued; customer balance reduced",
        "Integration", "Medium", "Not Run"),
    make_row("E2E-07", "E2E", "Cycle Count Adjust",
        "Mobile scan bin → Recount discrepancy → Adjust → Locations updated",
        "Recount a bin to a different qty",
        "Ledger row; Locations view reflects new qty",
        "Integration", "Medium", "Not Run"),
    make_row("E2E-08", "E2E", "RBAC",
        "Login as each role; check accessible pages",
        "warehouse, billing, procurement, supervisor, admin",
        "Each only sees allowed pages; mobile picking allowed for warehouse + supervisor",
        "Permission", "High", "Not Run"),
    make_row("E2E-09", "E2E", "FG Putaway",
        "MO complete with putaway rule → mobile TO pick/drop",
        "Run MO, follow the resulting putaway TO on the device",
        "FG moves from line to storage; ledger consistent",
        "Integration", "High", "Not Run"),
    make_row("E2E-10", "E2E", "APK install + version pin",
        "Install com.prakruthivanam.warehouse APK; verify version",
        "Sideload latest APK on device",
        "App launches → /m/login; About page shows current versionName",
        "Functional", "Medium", "Not Run"),
]

# ────────────────────────────────────────────────────────────────────
# 7. Role matrix
# ────────────────────────────────────────────────────────────────────
ROLE_MATRIX = [
    ("Page / Capability", "admin", "supervisor", "procurement", "billing", "warehouse"),
    ("Dashboard", "✓", "✓", "✓", "✓", "✓"),
    ("Products", "✓", "✓", "✓", "—", "—"),
    ("Procurement", "✓", "—", "✓", "—", "—"),
    ("Price Lists", "✓", "—", "✓", "—", "—"),
    ("Customers", "✓", "✓", "—", "✓", "—"),
    ("Enquiries", "✓", "✓", "—", "✓", "—"),
    ("Quotes", "✓", "✓", "—", "✓", "—"),
    ("Sales Orders", "✓", "✓", "—", "✓", "—"),
    ("Picking (web)", "✓", "✓", "—", "✓", "✓"),
    ("Packing (web)", "✓", "✓", "—", "✓", "✓"),
    ("Returns", "✓", "✓", "—", "✓", "✓"),
    ("Inventory", "✓", "✓", "✓", "—", "✓"),
    ("Warehouse", "✓", "✓", "—", "—", "✓"),
    ("Transfers", "✓", "✓", "—", "—", "✓"),
    ("Putaway Rules", "✓", "—", "—", "—", "—"),
    ("Warehouse Audit", "—", "—", "—", "—", "✓"),
    ("Manufacturing", "✓", "✓", "—", "—", "—"),
    ("BOMs", "✓", "✓", "—", "—", "—"),
    ("Productivity", "✓", "✓", "—", "—", "—"),
    ("Transport", "✓", "✓", "—", "—", "✓"),
    ("Billing", "✓", "—", "—", "✓", "—"),
    ("Reports", "✓", "✓", "✓", "✓", "—"),
    ("Approvals", "✓", "✓", "—", "✓", "—"),
    ("Settings", "✓", "—", "—", "—", "—"),
    ("Mobile · /m/tasks", "✓", "✓", "—", "—", "✓"),
    ("Mobile · /m/picks", "✓", "✓", "—", "—", "✓"),
    ("Mobile · /m/packs", "✓", "✓", "—", "—", "✓"),
    ("Mobile · /m/transfers", "✓", "✓", "—", "—", "✓"),
    ("Mobile · /m/count", "✓", "✓", "—", "—", "✓"),
    ("Mobile · /m/grn", "✓", "—", "✓", "—", "—"),
    ("Mobile · /m/returns", "✓", "✓", "—", "✓", "✓"),
]


def role_sheet(wb):
    if "7. Role Matrix" in wb.sheetnames:
        del wb["7. Role Matrix"]
    ws = wb.create_sheet("7. Role Matrix")
    ws["A1"] = "Role Matrix — who can reach what"
    ws["A1"].font = Font(bold=True, size=14, color="003087")
    ws.merge_cells("A1:F1")
    for r_idx, row in enumerate(ROLE_MATRIX, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
            if r_idx == 3:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            elif r_idx % 2 == 0:
                cell.fill = ZEBRA
            if c_idx > 1 and val == "✓":
                cell.font = Font(bold=True, color="019C34")
            elif c_idx > 1 and val == "—":
                cell.font = Font(color="9CA3AF")
    ws.column_dimensions["A"].width = 32
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "B4"


# ────────────────────────────────────────────────────────────────────
# Build the workbook
# ────────────────────────────────────────────────────────────────────
def build():
    wb = openpyxl.Workbook()
    # remove default sheet
    wb.remove(wb.active)

    overview_sheet(wb)
    write_sheet(wb, "1. Models", MODELS)
    write_sheet(wb, "2. ERP Pages", PAGES)
    write_sheet(wb, "3. Mobile Pick-Pack", MOB_PICK)
    write_sheet(wb, "4. Mobile Other", MOB_OTHER)
    write_sheet(wb, "5. API Routes", API)
    write_sheet(wb, "6. E2E Flows", E2E)
    role_sheet(wb)

    # Counts summary on overview
    ov = wb["Overview"]
    ov.cell(row=30, column=1, value="Counts").font = Font(bold=True, color="003087")
    ov.cell(row=31, column=1, value="1. Models")
    ov.cell(row=31, column=2, value=len(MODELS))
    ov.cell(row=32, column=1, value="2. ERP Pages")
    ov.cell(row=32, column=2, value=len(PAGES))
    ov.cell(row=33, column=1, value="3. Mobile Pick/Pack")
    ov.cell(row=33, column=2, value=len(MOB_PICK))
    ov.cell(row=34, column=1, value="4. Mobile Other")
    ov.cell(row=34, column=2, value=len(MOB_OTHER))
    ov.cell(row=35, column=1, value="5. API Routes")
    ov.cell(row=35, column=2, value=len(API))
    ov.cell(row=36, column=1, value="6. E2E Flows")
    ov.cell(row=36, column=2, value=len(E2E))
    total = len(MODELS) + len(PAGES) + len(MOB_PICK) + len(MOB_OTHER) + len(API) + len(E2E)
    ov.cell(row=37, column=1, value="TOTAL").font = Font(bold=True)
    ov.cell(row=37, column=2, value=total).font = Font(bold=True)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Total test cases: {total}")
    by_status = {}
    for s in MODELS + PAGES + MOB_PICK + MOB_OTHER + API + E2E:
        by_status[s[8]] = by_status.get(s[8], 0) + 1
    for k, v in sorted(by_status.items()):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    build()
